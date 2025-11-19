"""Excel export functionality with formatted sheets and analytics."""

import pandas as pd
from typing import List, Dict, Optional
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference


class ExcelExporter:
    """Export job applications and analytics to formatted Excel files."""

    def __init__(self, output_dir: Optional[Path] = None):
        """Initialize Excel exporter.

        Args:
            output_dir: Directory to save Excel files
        """
        if output_dir is None:
            from config import GENERATED_DOCS_DIR
            output_dir = GENERATED_DOCS_DIR / "exports"

        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_comprehensive_report(
        self,
        search_results: List[Dict],
        applications: List[Dict],
        projects: List[Dict],
        usage_stats: Dict[str, int],
        skill_counts: Dict[str, int],
        user_skills: List[str],
        user_name: str = "User",
        filename: Optional[str] = None
    ) -> Path:
        """Export comprehensive report with multiple sheets.

        Args:
            search_results: Job search results
            applications: Generated applications
            projects: User projects
            usage_stats: Project usage statistics
            skill_counts: Skill frequency counts
            user_skills: User's current skills
            user_name: User's name for report
            filename: Output filename (auto-generated if None)

        Returns:
            Path to generated Excel file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"job_search_report_{timestamp}.xlsx"

        output_path = self.output_dir / filename

        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Summary
            self._create_summary_sheet(
                writer,
                search_results,
                applications,
                projects,
                user_name
            )

            # Sheet 2: Search Results
            if search_results:
                self._create_search_results_sheet(writer, search_results)

            # Sheet 3: Applications
            if applications:
                self._create_applications_sheet(writer, applications)

            # Sheet 4: Project Performance
            if projects:
                self._create_project_performance_sheet(writer, projects, usage_stats)

            # Sheet 5: Skills Gap Analysis
            if skill_counts:
                total_jobs = len(search_results) if search_results else 0
                self._create_skills_gap_sheet(
                    writer,
                    skill_counts,
                    user_skills,
                    total_jobs
                )

        # Apply formatting
        self._apply_formatting(output_path)

        return output_path

    def _create_summary_sheet(
        self,
        writer,
        search_results: List[Dict],
        applications: List[Dict],
        projects: List[Dict],
        user_name: str
    ):
        """Create summary overview sheet."""
        summary_data = {
            'Metric': [
                'Report Generated',
                'User',
                '',
                'Total Jobs Found',
                'Total Applications',
                'Total Projects',
                '',
                'Average Match Score',
                'Top Match Score',
                'Applications This Week',
                '',
                'Top Matched Job',
                'Top Matched Company',
            ],
            'Value': [
                datetime.now().strftime("%Y-%m-%d %H:%M"),
                user_name,
                '',
                len(search_results),
                len(applications),
                len(projects),
                '',
                f"{sum(j.get('match_score', 0) for j in search_results) / len(search_results):.1f}%" if search_results else 'N/A',
                f"{max((j.get('match_score', 0) for j in search_results), default=0)}%" if search_results else 'N/A',
                len([a for a in applications if (datetime.utcnow() - a.get('generated_at', datetime.utcnow())).days < 7]),
                '',
                search_results[0].get('job_title', 'N/A') if search_results else 'N/A',
                search_results[0].get('company_name', 'N/A') if search_results else 'N/A',
            ]
        }

        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='Summary', index=False)

    def _create_search_results_sheet(self, writer, search_results: List[Dict]):
        """Create search results sheet."""
        rows = []
        for i, job in enumerate(search_results):
            row = {
                'Rank': i + 1,
                'Job Title': job.get('job_title', 'N/A'),
                'Company': job.get('company_name', 'N/A'),
                'Location': job.get('location', 'N/A'),
                'Match Score (%)': job.get('match_score', 0),
                'Semantic Similarity': round(job.get('semantic_similarity', 0) * 100, 1),
                'Job Type': job.get('job_type', 'N/A'),
                'Salary': job.get('salary', 'N/A'),
                'Posted': job.get('posted_date', 'N/A'),
                'Source': job.get('source', 'N/A'),
                'URL': job.get('url', 'N/A'),
                'Top Project': job.get('selected_projects', [{}])[0].get('title', 'N/A') if job.get('selected_projects') else 'N/A',
            }
            rows.append(row)

        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Job Search Results', index=False)

    def _create_applications_sheet(self, writer, applications: List[Dict]):
        """Create applications tracking sheet."""
        rows = []
        for i, app in enumerate(applications):
            job = app.get('job', {})
            cv = app.get('cv', {})
            metadata = cv.get('metadata', {})

            row = {
                'Application #': i + 1,
                'Job Title': job.get('job_title', 'N/A'),
                'Company': job.get('company_name', 'N/A'),
                'Location': job.get('location', 'N/A'),
                'Match Score (%)': job.get('match_score', 0),
                'CV Length (pages)': metadata.get('cv_length', 'N/A'),
                'Generation Method': metadata.get('generation_method', 'N/A'),
                'Word Count': metadata.get('word_count', 'N/A'),
                'Projects Included': len(metadata.get('projects_included', [])),
                'Generated At': app.get('generated_at', 'N/A'),
                'Status': 'Generated',
            }
            rows.append(row)

        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Applications', index=False)

    def _create_project_performance_sheet(
        self,
        writer,
        projects: List[Dict],
        usage_stats: Dict[str, int]
    ):
        """Create project performance sheet."""
        rows = []
        for project in projects:
            project_id = str(project.get('id', ''))
            usage_count = usage_stats.get(project_id, 0)

            row = {
                'Project Title': project.get('title', 'N/A'),
                'Technologies': ', '.join(project.get('technologies', [])),
                'Times Used': usage_count,
                'Usage %': 0,  # Will be calculated
                'GitHub URL': project.get('github_url', 'N/A'),
            }
            rows.append(row)

        df = pd.DataFrame(rows)

        # Calculate usage percentage
        total_usage = df['Times Used'].sum()
        if total_usage > 0:
            df['Usage %'] = (df['Times Used'] / total_usage * 100).round(1)

        # Sort by usage
        df = df.sort_values('Times Used', ascending=False)

        df.to_excel(writer, sheet_name='Project Performance', index=False)

    def _create_skills_gap_sheet(
        self,
        writer,
        skill_counts: Dict[str, int],
        user_skills: List[str],
        total_jobs: int
    ):
        """Create skills gap analysis sheet."""
        user_skills_lower = set(s.lower() for s in user_skills)
        rows = []

        for skill, count in sorted(skill_counts.items(), key=lambda x: x[1], reverse=True)[:30]:
            has_skill = skill.lower() in user_skills_lower
            percentage = (count / total_jobs * 100) if total_jobs > 0 else 0

            row = {
                'Skill': skill,
                'Jobs Requiring': count,
                '% of Jobs': round(percentage, 1),
                'You Have': 'Yes' if has_skill else 'No',
                'Gap': 'Learn' if not has_skill else 'Have',
                'Priority': 'High' if not has_skill and percentage >= 50 else 'Medium' if not has_skill and percentage >= 25 else 'Low'
            }
            rows.append(row)

        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Skills Gap Analysis', index=False)

    def _apply_formatting(self, filepath: Path):
        """Apply formatting to Excel file."""
        try:
            workbook = load_workbook(filepath)

            # Format each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Header formatting
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=12)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Apply header formatting
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width

                # Freeze header row
                sheet.freeze_panes = 'A2'

            # Add conditional formatting for match scores in Search Results
            if 'Job Search Results' in workbook.sheetnames:
                sheet = workbook['Job Search Results']

                # Find match score column
                for col_idx, cell in enumerate(sheet[1], start=1):
                    if 'Match Score' in str(cell.value):
                        # Apply color scale to match scores
                        from openpyxl.formatting.rule import ColorScaleRule

                        color_scale = ColorScaleRule(
                            start_type='num', start_value=0, start_color='F8696B',
                            mid_type='num', mid_value=50, mid_color='FFEB84',
                            end_type='num', end_value=100, end_color='63BE7B'
                        )

                        col_letter = sheet.cell(row=1, column=col_idx).column_letter
                        sheet.conditional_formatting.add(
                            f'{col_letter}2:{col_letter}{sheet.max_row}',
                            color_scale
                        )

            workbook.save(filepath)

        except Exception as e:
            print(f"Warning: Could not apply formatting: {e}")

    def export_simple_excel(
        self,
        data: List[Dict],
        sheet_name: str = "Data",
        filename: Optional[str] = None
    ) -> Path:
        """Export simple data to Excel with basic formatting.

        Args:
            data: List of dictionaries to export
            sheet_name: Name of the sheet
            filename: Output filename (auto-generated if None)

        Returns:
            Path to generated Excel file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{timestamp}.xlsx"

        output_path = self.output_dir / filename

        df = pd.DataFrame(data)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Apply basic formatting
        self._apply_formatting(output_path)

        return output_path


# Global exporter instance
excel_exporter = ExcelExporter()
